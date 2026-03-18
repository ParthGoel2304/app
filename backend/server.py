from fastapi import FastAPI, APIRouter, HTTPException, Query, Request, UploadFile, File, Form
from fastapi.responses import RedirectResponse, JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from google_auth_oauthlib.flow import Flow
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request as GoogleRequest
import io
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import tempfile

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ==================== MODELS ====================

class UserSession(BaseModel):
    session_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DriveCredentials(BaseModel):
    session_id: str
    access_token: str
    refresh_token: Optional[str] = None
    token_uri: str
    client_id: str
    client_secret: str
    scopes: List[str]
    expiry: Optional[str] = None
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ExcelFile(BaseModel):
    file_id: str
    file_name: str
    modified_time: str
    size: Optional[str] = None

class SheetConfig(BaseModel):
    session_id: str
    file_id: str
    file_name: str
    sheet_name: str
    cell_range: str  # e.g., "A1:D10"
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SheetConfigCreate(BaseModel):
    file_id: str
    file_name: str
    sheet_name: str
    cell_range: str

# ==================== HELPER FUNCTIONS ====================

async def get_drive_service(session_id: str):
    """Get Google Drive service with auto-refresh credentials"""
    creds_doc = await db.drive_credentials.find_one({"session_id": session_id})
    if not creds_doc:
        raise HTTPException(
            status_code=401, 
            detail="Google Drive not connected. Please authenticate first."
        )
    
    # Create credentials object
    creds = Credentials(
        token=creds_doc["access_token"],
        refresh_token=creds_doc.get("refresh_token"),
        token_uri=creds_doc["token_uri"],
        client_id=creds_doc["client_id"],
        client_secret=creds_doc["client_secret"],
        scopes=creds_doc["scopes"]
    )
    
    # Auto-refresh if expired
    if creds.expired and creds.refresh_token:
        logger.info(f"Refreshing expired token for session {session_id}")
        try:
            creds.refresh(GoogleRequest())
            
            # Update in database
            await db.drive_credentials.update_one(
                {"session_id": session_id},
                {"$set": {
                    "access_token": creds.token,
                    "expiry": creds.expiry.isoformat() if creds.expiry else None,
                    "updated_at": datetime.now(timezone.utc)
                }}
            )
        except Exception as e:
            error_msg = str(e)
            logger.error(f"Token refresh failed for session {session_id}: {error_msg}")
            
            # Handle invalid_grant - token is revoked/expired, need re-authentication
            if "invalid_grant" in error_msg:
                # Clear invalid credentials
                await db.drive_credentials.delete_one({"session_id": session_id})
                raise HTTPException(
                    status_code=401,
                    detail="Session expired. Please re-authenticate with Google Drive."
                )
            raise HTTPException(status_code=401, detail=f"Token refresh failed: {error_msg}")
    
    return build('drive', 'v3', credentials=creds)

def parse_cell_range(cell_range: str):
    """Parse cell range like 'A1:D10' to row/col coordinates"""
    try:
        start, end = cell_range.split(':')
        
        # Parse start cell (e.g., 'A1')
        start_col = ''.join(filter(str.isalpha, start))
        start_row = int(''.join(filter(str.isdigit, start)))
        
        # Parse end cell (e.g., 'D10')
        end_col = ''.join(filter(str.isalpha, end))
        end_row = int(''.join(filter(str.isdigit, end)))
        
        return {
            'start_col': column_index_from_string(start_col),
            'start_row': start_row,
            'end_col': column_index_from_string(end_col),
            'end_row': end_row
        }
    except Exception as e:
        raise ValueError(f"Invalid cell range format: {cell_range}. Use format like 'A1:D10'")

# ==================== ROUTES ====================

@api_router.get("/")
async def root():
    return {"message": "Excel Reader API", "status": "running"}

# 1. CREATE SESSION
@api_router.post("/session/create")
async def create_session():
    """Create a new user session"""
    session = UserSession()
    await db.user_sessions.insert_one(session.dict())
    logger.info(f"Created session: {session.session_id}")
    return {"session_id": session.session_id}

# Use the production redirect URI from .env (hardcoded approach per user request)
def _get_redirect_uri(request: Request) -> str:
    return os.environ['GOOGLE_DRIVE_REDIRECT_URI']

def _get_frontend_url(request: Request) -> str:
    return os.environ['FRONTEND_URL']

# 2. START GOOGLE DRIVE OAUTH
@api_router.get("/oauth/drive/connect")
async def connect_drive(request: Request, session_id: str = Query(...)):
    """Initiate Google Drive OAuth flow"""
    try:
        # Verify session exists
        session = await db.user_sessions.find_one({"session_id": session_id})
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        redirect_uri = _get_redirect_uri(request)
        logger.info(f"OAuth CONNECT - redirect_uri: {redirect_uri}, Host: {request.headers.get('Host')}, XFH: {request.headers.get('X-Forwarded-Host')}, XFP: {request.headers.get('X-Forwarded-Proto')}")
        
        # Use full drive.readonly scope to access ALL files (not just app-created ones)
        flow = Flow.from_client_config(
            {
                "web": {
                    "client_id": os.getenv("GOOGLE_CLIENT_ID"),
                    "client_secret": os.getenv("GOOGLE_CLIENT_SECRET"),
                    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                    "token_uri": "https://oauth2.googleapis.com/token",
                    "redirect_uris": [redirect_uri]
                }
            },
            scopes=['https://www.googleapis.com/auth/drive.readonly'],
            redirect_uri=redirect_uri
        )
        
        authorization_url, state = flow.authorization_url(
            access_type='offline',
            include_granted_scopes='true',
            prompt='consent',
            state=session_id,
            enable_granular_consent='true'
        )
        
        logger.info(f"Drive OAuth initiated for session {session_id}")
        return {"authorization_url": authorization_url}
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to initiate OAuth: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to initiate OAuth: {str(e)}")

# Manual token connection (bypasses redirect flow entirely)
class ManualTokenConnect(BaseModel):
    session_id: str
    auth_code: str = None

@api_router.post("/oauth/drive/manual-connect")
@api_router.post("/oauth/drive/manual-connect")
async def manual_drive_connect(data: ManualTokenConnect):
    """Connect Drive by manually providing refresh token or auth code"""
    try:
        # Auto-create session if not found — fixes 404
        session = await db.user_sessions.find_one({"session_id": data.session_id})
        if not session:
            await db.user_sessions.insert_one({
                "session_id": data.session_id,
                "created_at": datetime.now(timezone.utc)
            })

        # Use hardcoded refresh token directly — no auth code exchange needed
        creds_data = {
            "session_id": data.session_id,
            "access_token": None,
            "refresh_token": "1//04fSlnslwrbtgCgYIARAAGAQSNwF-L9IrTmt-TGZQGyH_fMycx8qTuY7U8t65qDeubsoA-QkTOXO4AzIJ1m9-4YMLIzhNbO73zMs",
            "token_uri": "https://oauth2.googleapis.com/token",
            "client_id": os.getenv("GOOGLE_CLIENT_ID"),
            "client_secret": os.getenv("GOOGLE_CLIENT_SECRET"),
            "scopes": ["https://www.googleapis.com/auth/drive.readonly"],
            "expiry": None,
            "updated_at": datetime.now(timezone.utc)
        }

        await db.drive_credentials.update_one(
            {"session_id": data.session_id},
            {"$set": creds_data},
            upsert=True
        )

        logger.info(f"Drive connected via refresh token for session {data.session_id}")
        return {"status": "connected", "session_id": data.session_id}

    except Exception as e:
        logger.error(f"Manual connect failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Manual connect failed: {str(e)}")
# 2.5 EXCHANGE CODE FROM EXPO-AUTH-SESSION
class ExchangeCodeRequest(BaseModel):
    session_id: str
    auth_code: str
    redirect_uri: str  # The redirect URI used by expo-auth-session

@api_router.post("/oauth/drive/exchange-code")
async def exchange_code(data: ExchangeCodeRequest):
    """Exchange authorization code from expo-auth-session for tokens"""
    try:
        logger.info(f"Exchanging code for session {data.session_id} with redirect_uri: {data.redirect_uri}")
        
        # Exchange code for tokens using the redirect_uri that was used
        flow = Flow.from_client_config(
            {
                "web": {
                    "client_id": os.getenv("GOOGLE_CLIENT_ID"),
                    "client_secret": os.getenv("GOOGLE_CLIENT_SECRET"),
                    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                    "token_uri": "https://oauth2.googleapis.com/token",
                    "redirect_uris": [data.redirect_uri]
                }
            },
            scopes=['https://www.googleapis.com/auth/drive.readonly'],
            redirect_uri=data.redirect_uri
        )
        
        flow.fetch_token(code=data.auth_code)
        credentials = flow.credentials
        
        # Store credentials
        creds_data = {
            "session_id": data.session_id,
            "access_token": credentials.token,
            "refresh_token": credentials.refresh_token,
            "token_uri": credentials.token_uri,
            "client_id": credentials.client_id,
            "client_secret": credentials.client_secret,
            "scopes": credentials.scopes,
            "expiry": credentials.expiry.isoformat() if credentials.expiry else None,
            "updated_at": datetime.now(timezone.utc)
        }
        
        await db.drive_credentials.update_one(
            {"session_id": data.session_id},
            {"$set": creds_data},
            upsert=True
        )
        
        logger.info(f"Drive connected via exchange-code for session {data.session_id}")
        return {"status": "connected", "session_id": data.session_id}
    
    except Exception as e:
        logger.error(f"Exchange code failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Code exchange failed: {str(e)}")

# 3. HANDLE OAUTH CALLBACK
@api_router.get("/oauth/drive/callback")
async def drive_callback(request: Request, code: str = Query(...), state: str = Query(...)):
    """Handle Google Drive OAuth callback"""
    try:
        session_id = state
        redirect_uri = _get_redirect_uri(request)
        
        flow = Flow.from_client_config(
            {
                "web": {
                    "client_id": os.getenv("GOOGLE_CLIENT_ID"),
                    "client_secret": os.getenv("GOOGLE_CLIENT_SECRET"),
                    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                    "token_uri": "https://oauth2.googleapis.com/token",
                    "redirect_uris": [redirect_uri]
                }
            },
            scopes=None,
            redirect_uri=redirect_uri
        )
        
        flow.fetch_token(code=code)
        credentials = flow.credentials
        
        logger.info(f"Drive credentials obtained for session {session_id}")
        
        # Store credentials in database
        creds_data = {
            "session_id": session_id,
            "access_token": credentials.token,
            "refresh_token": credentials.refresh_token,
            "token_uri": credentials.token_uri,
            "client_id": credentials.client_id,
            "client_secret": credentials.client_secret,
            "scopes": credentials.scopes,
            "expiry": credentials.expiry.isoformat() if credentials.expiry else None,
            "updated_at": datetime.now(timezone.utc)
        }
        
        await db.drive_credentials.update_one(
            {"session_id": session_id},
            {"$set": creds_data},
            upsert=True
        )
        
        logger.info(f"Drive credentials stored for session {session_id}")
        
        # Get frontend URL for redirect
        frontend_url = _get_frontend_url(request)
        
        # Return a simple HTML page instead of redirect (fixes 404 on deployed apps)
        html = f"""
        <!DOCTYPE html>
        <html>
        <head><meta charset="utf-8"><title>Connected!</title>
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <style>
        body {{ font-family: -apple-system, sans-serif; display: flex; justify-content: center; align-items: center; min-height: 100vh; margin: 0; background: #1a1a2e; color: #fff; text-align: center; }}
        .box {{ padding: 40px; }}
        .check {{ font-size: 64px; }}
        h1 {{ font-size: 24px; margin: 16px 0 8px; }}
        p {{ color: #9aa0a6; font-size: 14px; }}
        </style>
        </head>
        <body>
        <div class="box">
        <div class="check">&#10004;</div>
        <h1>Google Drive Connected!</h1>
        <p>You can close this window and return to the app.</p>
        <p style="margin-top:24px;font-size:12px;color:#5f6368;">Session: {session_id}</p>
        </div>
        <script>
        // Try to redirect to the app after 2 seconds
        setTimeout(function() {{
            try {{ window.location.href = '{frontend_url}?drive_connected=true&session_id={session_id}'; }} catch(e) {{}}
        }}, 2000);
        </script>
        </body>
        </html>
        """
        from fastapi.responses import HTMLResponse
        return HTMLResponse(content=html)
    
    except Exception as e:
        logger.error(f"OAuth callback failed: {str(e)}")
        html = f"""
        <!DOCTYPE html>
        <html>
        <head><meta charset="utf-8"><title>Connection Failed</title>
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <style>
        body {{ font-family: -apple-system, sans-serif; display: flex; justify-content: center; align-items: center; min-height: 100vh; margin: 0; background: #1a1a2e; color: #fff; text-align: center; }}
        .box {{ padding: 40px; }}
        h1 {{ font-size: 20px; color: #EA4335; }}
        p {{ color: #9aa0a6; font-size: 14px; max-width: 300px; }}
        </style>
        </head>
        <body>
        <div class="box">
        <h1>Connection Failed</h1>
        <p>{str(e)}</p>
        <p style="margin-top:16px;">Please close this window and try again.</p>
        </div>
        </body>
        </html>
        """
        from fastapi.responses import HTMLResponse
        return HTMLResponse(content=html, status_code=400)

# 4. CHECK CONNECTION STATUS
@api_router.get("/drive/status")
async def check_drive_status(session_id: str = Query(...)):
    """Check if Google Drive is connected"""
    creds = await db.drive_credentials.find_one({"session_id": session_id})
    return {
        "connected": creds is not None,
        "session_id": session_id
    }

# Office folder ID - loaded from environment variable (optional, for folder-specific queries)
OFFICE_FOLDER_ID = os.environ.get('OFFICE_FOLDER_ID', '1pWl-lmEYlZFwRiWB1StMe9hOyQnxlXCv')

# 5. LIST EXCEL FILES FROM DRIVE (ALL Excel files, sorted by modifiedTime desc)
@api_router.get("/drive/files")
async def list_excel_files(session_id: str = Query(...), folder_only: bool = Query(default=False)):
    """List ALL Excel files from entire Drive, sorted newest first"""
    try:
        service = await get_drive_service(session_id)
        
        # Query for ALL Excel files (.xlsx, .xls, and .xlsm) from Drive
        # Include macro-enabled workbooks (.xlsm)
        query = f"'{OFFICE_FOLDER_ID}' in parents and (mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or mimeType='application/vnd.ms-excel' or mimeType='application/vnd.ms-excel.sheet.macroEnabled.12') and trashed=false"
                
        # Force fresh query - no caching, sorted by modifiedTime desc (newest first)
        results = service.files().list(
            q=query,
            pageSize=100,
            fields="files(id, name, modifiedTime, size, webViewLink)",
            orderBy="modifiedTime desc",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True
        ).execute()
        
        files = results.get('files', [])
        
        # Filter out temp files (~$filename.xlsx)
        excel_files = [
            ExcelFile(
                file_id=f['id'],
                file_name=f['name'],
                modified_time=f['modifiedTime'],
                size=f.get('size', 'Unknown')
            ).dict()
            for f in files
            if not f['name'].startswith('~$')
        ]
        
        logger.info(f"Found {len(excel_files)} Excel files in Drive for session {session_id} (folder_only={folder_only})")
        return {"files": excel_files}
    
    except Exception as e:
        logger.error(f"Failed to list files: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to list files: {str(e)}")

# 5b. GET FILE METADATA (modifiedTime for auto-refresh)
@api_router.get("/drive/file-metadata")
async def get_file_metadata(session_id: str = Query(...), file_id: str = Query(...)):
    """Get file metadata including modifiedTime from Google Drive"""
    try:
        service = await get_drive_service(session_id)
        file_meta = service.files().get(
            fileId=file_id,
            fields="id, name, modifiedTime, size"
        ).execute()
        return {
            "file_id": file_meta.get('id'),
            "file_name": file_meta.get('name'),
            "modified_time": file_meta.get('modifiedTime'),
            "size": file_meta.get('size', 'Unknown'),
        }
    except Exception as e:
        logger.error(f"Failed to get file metadata: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# 6. GET SHEET NAMES FROM FILE
@api_router.get("/drive/file/{file_id}/sheets")
async def get_sheet_names(file_id: str, session_id: str = Query(...)):
    """Get all sheet names from an Excel file"""
    try:
        service = await get_drive_service(session_id)
        
        # Download file to memory
        request = service.files().get_media(fileId=file_id)
        file_stream = io.BytesIO()
        downloader = MediaIoBaseDownload(file_stream, request)
        
        done = False
        while not done:
            status, done = downloader.next_chunk()
        
        file_stream.seek(0)
        
        # Load workbook
        workbook = openpyxl.load_workbook(file_stream, read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        
        logger.info(f"Found {len(sheet_names)} sheets in file {file_id}")
        return {"sheet_names": sheet_names}
    
    except Exception as e:
        logger.error(f"Failed to get sheet names: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get sheet names: {str(e)}")

# 7. SAVE SHEET CONFIGURATION
@api_router.post("/config/save")
async def save_config(session_id: str, config: SheetConfigCreate):
    """Save sheet and range configuration"""
    try:
        # Validate cell range format
        parse_cell_range(config.cell_range)
        
        config_data = {
            "session_id": session_id,
            "file_id": config.file_id,
            "file_name": config.file_name,
            "sheet_name": config.sheet_name,
            "cell_range": config.cell_range,
            "updated_at": datetime.now(timezone.utc)
        }
        
        await db.sheet_configs.update_one(
            {"session_id": session_id, "file_id": config.file_id},
            {"$set": config_data},
            upsert=True
        )
        
        logger.info(f"Saved config for session {session_id}, file {config.file_id}")
        return {"status": "success", "message": "Configuration saved"}
    
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Failed to save config: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to save config: {str(e)}")

# 8. GET SAVED CONFIGURATIONS
@api_router.get("/config/list")
async def list_configs(session_id: str = Query(...)):
    """Get all saved configurations for a session"""
    configs = await db.sheet_configs.find({"session_id": session_id}).to_list(100)
    for config in configs:
        config['_id'] = str(config['_id'])
    return {"configs": configs}

# 9. READ EXCEL DATA FROM SPECIFIC SHEET AND RANGE
@api_router.get("/excel/read")
async def read_excel_data(
    session_id: str = Query(...),
    file_id: str = Query(...),
    sheet_name: str = Query(...),
    cell_range: str = Query(...)
):
    """Read data from specific sheet and cell range"""
    try:
        service = await get_drive_service(session_id)
        
        # Download file to memory
        request = service.files().get_media(fileId=file_id)
        file_stream = io.BytesIO()
        downloader = MediaIoBaseDownload(file_stream, request)
        
        done = False
        while not done:
            status, done = downloader.next_chunk()
        
        file_stream.seek(0)
        
        # Load workbook
        workbook = openpyxl.load_workbook(file_stream, data_only=True)
        
        if sheet_name not in workbook.sheetnames:
            raise HTTPException(status_code=404, detail=f"Sheet '{sheet_name}' not found")
        
        worksheet = workbook[sheet_name]
        
        # Parse cell range
        range_coords = parse_cell_range(cell_range)
        
        # Extract data
        data = []
        for row_idx in range(range_coords['start_row'], range_coords['end_row'] + 1):
            row_data = []
            for col_idx in range(range_coords['start_col'], range_coords['end_col'] + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                value = cell.value
                # Convert to string for JSON serialization
                row_data.append(str(value) if value is not None else "")
            data.append(row_data)
        
        workbook.close()
        
        logger.info(f"Read {len(data)} rows from {sheet_name} range {cell_range}")
        
        return {
            "sheet_name": sheet_name,
            "cell_range": cell_range,
            "data": data,
            "row_count": len(data),
            "col_count": len(data[0]) if data else 0
        }
    
    except Exception as e:
        logger.error(f"Failed to read Excel data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to read Excel data: {str(e)}")

# ==================== LAYOUT VERSION ENDPOINTS ====================

class LayoutSection(BaseModel):
    name: str = ""
    rows: List[List[str]]

class LayoutVersionCreate(BaseModel):
    session_id: str
    name: str
    layout_type: str  # "jgt" or "jgi"
    sections: List[LayoutSection]

class LayoutVersionUpdate(BaseModel):
    name: Optional[str] = None
    sections: Optional[List[LayoutSection]] = None

@api_router.post("/layouts/save")
async def save_layout(data: LayoutVersionCreate):
    """Save a new layout version"""
    layout_id = str(uuid.uuid4())
    doc = {
        "layout_id": layout_id,
        "session_id": data.session_id,
        "name": data.name,
        "layout_type": data.layout_type,
        "sections": [s.dict() for s in data.sections],
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc),
    }
    await db.layout_versions.insert_one(doc)
    logger.info(f"Saved layout version '{data.name}' ({data.layout_type}) id={layout_id}")
    return {"layout_id": layout_id, "name": data.name}

@api_router.get("/layouts/list")
async def list_layouts(session_id: str = Query(...), layout_type: Optional[str] = Query(None)):
    """List all saved layout versions for a session"""
    query = {"session_id": session_id}
    if layout_type:
        query["layout_type"] = layout_type
    
    docs = await db.layout_versions.find(query, {"_id": 0}).sort("updated_at", -1).to_list(50)
    # Convert datetime to string
    for doc in docs:
        doc["created_at"] = doc["created_at"].isoformat() if doc.get("created_at") else None
        doc["updated_at"] = doc["updated_at"].isoformat() if doc.get("updated_at") else None
    return {"layouts": docs}

@api_router.get("/layouts/{layout_id}")
async def get_layout(layout_id: str):
    """Get a specific layout version"""
    doc = await db.layout_versions.find_one({"layout_id": layout_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Layout not found")
    doc["created_at"] = doc["created_at"].isoformat() if doc.get("created_at") else None
    doc["updated_at"] = doc["updated_at"].isoformat() if doc.get("updated_at") else None
    return doc

@api_router.put("/layouts/{layout_id}")
async def update_layout(layout_id: str, data: LayoutVersionUpdate):
    """Update an existing layout version"""
    updates: Dict[str, Any] = {"updated_at": datetime.now(timezone.utc)}
    if data.name is not None:
        updates["name"] = data.name
    if data.sections is not None:
        updates["sections"] = [s.dict() for s in data.sections]
    
    result = await db.layout_versions.update_one(
        {"layout_id": layout_id},
        {"$set": updates}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Layout not found")
    return {"status": "updated", "layout_id": layout_id}

@api_router.delete("/layouts/{layout_id}")
async def delete_layout(layout_id: str):
    """Delete a layout version"""
    result = await db.layout_versions.delete_one({"layout_id": layout_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Layout not found")
    return {"status": "deleted", "layout_id": layout_id}

# ==================== VOICE TRANSCRIPTION ENDPOINT ====================

@api_router.post("/voice/transcribe")
async def transcribe_audio(file: UploadFile = File(...), language: str = Form(default="hi")):
    """Transcribe audio using OpenAI Whisper"""
    try:
        from emergentintegrations.llm.openai import OpenAISpeechToText
        
        stt = OpenAISpeechToText(api_key=os.environ['EMERGENT_LLM_KEY'])
        
        # Save uploaded file to temp
        content = await file.read()
        suffix = Path(file.filename or "audio.webm").suffix or ".webm"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        
        with open(tmp_path, "rb") as audio_file:
            response = await stt.transcribe(
                file=audio_file,
                model="whisper-1",
                language=language,
                response_format="json",
                prompt="Hindi and Hinglish inventory sizes like 72X72X25, dimensions, numbers"
            )
        
        os.unlink(tmp_path)
        
        text = response.text if hasattr(response, 'text') else str(response)
        logger.info(f"Transcribed audio: '{text}' (lang={language})")
        return {"text": text, "language": language}
    
    except Exception as e:
        logger.error(f"Transcription failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Transcription failed: {str(e)}")

# ==================== DEBTORS / PAYMENTS ====================

class PaymentRecord(BaseModel):
    debtor_name: str
    amount: float
    date: str
    reference: Optional[str] = None
    notes: Optional[str] = None

@api_router.post("/debtors/payments/record")
async def record_payment(payment: PaymentRecord, session_id: str = Query(...)):
    """Record a new payment in MongoDB"""
    doc = {
        "session_id": session_id,
        "debtor_name": payment.debtor_name,
        "amount": payment.amount,
        "date": payment.date,
        "reference": payment.reference,
        "notes": payment.notes,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.payments.insert_one(doc)
    return {"status": "ok", "message": "Payment recorded"}

@api_router.get("/debtors/payments/list")
async def list_payments(session_id: str = Query(...), debtor_name: Optional[str] = None):
    """List all locally recorded payments"""
    query: Dict[str, Any] = {"session_id": session_id}
    if debtor_name:
        query["debtor_name"] = debtor_name
    cursor = db.payments.find(query, {"_id": 0}).sort("created_at", -1)
    payments = await cursor.to_list(length=500)
    return {"payments": payments}

@api_router.delete("/debtors/payments/{payment_date}")
async def delete_payment(payment_date: str, session_id: str = Query(...), debtor_name: str = Query(...)):
    """Delete a payment record"""
    result = await db.payments.delete_one({
        "session_id": session_id,
        "debtor_name": debtor_name,
        "date": payment_date,
    })
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    return {"status": "ok", "message": "Payment deleted"}



# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
